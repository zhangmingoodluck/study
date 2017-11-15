#coding:utf-8
from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook("D:/password.xlsx")
print(wb.sheetnames)
# 显示表名，表行数，表列数
print("Work Sheet Titile: " + wb["blacklist"].title)
print("Work Sheet Rows:", wb["blacklist"].min_row)
print("Work Sheet Cols: ", wb["blacklist"].min_column)
# 给一个单元格赋值
wb["blacklist"]['C1'] = "42"

# 在最后添加一行数据，用,号间隔blacklist
wb["blacklist"].append([1, 2, 3])

#保存
wb.save("D:/password.xlsx")