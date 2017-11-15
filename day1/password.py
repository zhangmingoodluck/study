#Author Mike
#readme
#从本地文件读取用户名和密码
#验证用户名和密码
#3次失败锁定用户名
#import xlrd
import openpyxl
from openpyxl import load_workbook
count=0
filename="d:/password.xlsx"
sheet_userinfo="userinfo"
sheet_blacklist="blacklist"
workbook= openpyxl.load_workbook(filename) #打开excel文件
#sheet_user=workbook.get_sheet_by_name(sheet_userinfo)#获取userinfoexcel表
try:
    sheet_user = workbook.get_sheet_by_name(sheet_userinfo)
except:
    print ("no sheet in %s named " % sheet_userinfo )
#获取blacklist表
#sheet_black=workbook.get_sheet_by_name(sheet_blacklist)
try:
    sheet_black = workbook.get_sheet_by_name(sheet_blacklist)
except:
    print ("no sheet in %s named " % sheet_blacklist )
black_num_of_rows=sheet_black.max_row #获取黑名单最大行数
#print(black_num_of_rows)
user_num_of_rows= sheet_user.max_row#获取用户名单最大行数
#print(user_num_of_rows)
count=0
stop=True
#先判断用户名是否在黑名单里面

while stop :
# 获取用户名和密码
    getuser = input("please input username:")  # 获取用户名
    getpassword = input("please input password:")  # 获取用户密码
    row_black=2
#黑名单校验，如果用户信息在黑名单里，则提示用户账户已被锁定

    while row_black<=black_num_of_rows:
#       print(row)
        user = str(sheet_black['A' + str(row_black)].value)
#        print(user)
#        print(getuser)
        if user==getuser:
            print ("your account have been locked")
            stop=False
            break
        row_black+=1
#用户名校验，非黑名单用户，3次密码失败则锁定账户
    row_user=2
    while row_user<=user_num_of_rows and stop==True:
        user = str(sheet_user['A' + str(row_user)].value)
#        print(row_user)
#        print(user)
#        print(getuser)
        password = str(sheet_user['B' + str(row_user)].value)
#        print(password)
        if user==getuser and password==getpassword:
            print("welcome")
            stop=False
            break
        else:
            row_user += 1
    if stop==True:#密码和账户校验失败提示错误
        print("wrong username or password")
    count += 1
    if count==3:#密码和账户校验失败3次提示进入黑名单
        stop=False
        print("You have tried 3 times,goodluck next time")
        #excel写入操作
        wb = load_workbook("D:/password.xlsx")
        # 在最后添加一行数据，用,号间隔blacklist
        wb["blacklist"].append([getuser])

        # 保存，黑名单结果
        wb.save("D:/password.xlsx")

