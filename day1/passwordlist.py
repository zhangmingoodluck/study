#Author Mike
#readme
#从本地文件读取用户名和密码
#验证用户名和密码
#3次失败锁定用户名
#---------------------------我是分割线----------------------
from  datetime  import  *                                                                                              #加载日期处理包
from openpyxl import load_workbook                                                                                     #加载excel的xlsx的读写方法,xls文件需要另外的包
filename="d:/password.xlsx"                                                                                           #定义excel的文件路径和文件名
sheet_userinfo="userinfo"                                                                                              #定义用户，密码表的表名
sheet_blacklist="blacklist"                                                                                            #定义黑名单表名
sheet_loglist="log"                                                                                                     #定义登录日志表表名
workbook= load_workbook(filename)                                                                                        #打开excel文件
try:                                                                                                                    #获取登录log表，
    sheet_log = workbook.get_sheet_by_name(sheet_loglist)
except:                                                                                                                #如果没有找到该日志表则提示，找不到这个表
    print ("no sheet in %s named " % sheet_loglist )
log_num_of_rows=sheet_log.max_row                                                                                        #获取log最大行数用于循环的最大次数
loglist=[]                                                                                                               #定义日志数组，用于容纳日志表读出的数据和增加后续的日志，用于统计当日的登录次数
i = datetime.now()                                                                                                       #获取当前系统日期
year=str(i.year)                                                                                                         #获取当前的年份
month=str(i.month)                                                                                                       #获取当前的月份
day=str(i.day)                                                                                                           #获取当前的日期
today=year + "/" + month +"/"+ day                                                                                      #拼出"YYYY/MM/DD"格式的日期
now=str(i)                                                                                                               #获取当前系统的时间
for i in range(2,log_num_of_rows+1):                                                                                    #for循环，从第二行开始到log的表的结束
    if str(sheet_log['B' + str(i)].value)==today and str(sheet_log['D' + str(i)].value)=="fail":                      #如果log记录日期是当日，且登录状态为失败
        loglist.append(str(sheet_log['A' + str(i)].value))                                                              #进入登录列表
try:                                                                                                                    # 获取userinfoexcel表
    sheet_user = workbook.get_sheet_by_name(sheet_userinfo)
except:                                                                                                                 #如果没有找到该日志表则提示，找不到这个表
    print ("no sheet in %s named " % sheet_userinfo )
user_num_of_rows= sheet_user.max_row                                                                                     #获取用户名单最大行数，用于循环的最大次数
userpasswordlist = []                                                                                                    #建立用户密码列表数组
userlist=[]                                                                                                              #建立用户名的列表
for i in range(2, user_num_of_rows+1):                                                                                  #加载excel 的用户名和密码到数组内，用于匹配用户名和密码
    userpasswordlist.append( str(sheet_user['A' + str(i)].value)+ str(sheet_user['B' + str(i)].value))                  #用户名和密码连接后放进数组用于匹配用户名密码，确定密码是否正确
    userlist.append(str(sheet_user['A' + str(i)].value))                                                                #用户名单独放一个数组，用于确定用户名是否正确
try:                                                                                                                    #获取blacklist表，用于确定登录的用户是否在黑名单上
    sheet_black = workbook.get_sheet_by_name(sheet_blacklist)
except:                                                                                                                 #如果找不到表就报错
    print ("no sheet in %s named " % sheet_blacklist )
black_num_of_rows=sheet_black.max_row                                                                                    #获取黑名单最大行数，用于加载数组时做循环的上限
blackuserlist=[]                                                                                                         #创建黑名单数组
for i in range(2,black_num_of_rows+1):                                                                                  #加载excel的黑名单列表到数组内
    blackuserlist.append(str(sheet_black['A' + str(i)].value))                                                          #用户名加入黑名单数据，方便后续匹配
count=0                                                                                                                  #控制log的行，每次循环增加一行日志
stop=True                                                                                                               #循环控制变量
while stop :                                                                                                            #开始循环
    getuser = input("please input username:")                                                                        # 获取用户名
    getpassword = input("please input password:")                                                                    # 获取用户密码
    getuserpassword=getuser + getpassword                                                                                #为了比对更方便
    count+=1                                                                                                             #增加登录日志
    if getuser in blackuserlist:                                                                                         #匹配用户名是否在黑名单列表中，如果在则提示用户名已被锁定退出循环
        print ("your account have been locked")                                                                     #提示用户，用户名是黑名单
        stop = False                                                                                                    #循环控制变量变为false
        break                                                                                                           #跳出循环
    if getuser in userlist:                                                                                             # 用户名不在黑名单中，且，用户名在用户列表中增加登录日志,进入下一层判断密码是否正确
        if getuserpassword in userpasswordlist:                                                                         # 匹配用户名和密码，成功则提示登录成功，跳出循环
            sheet_log['A' + str(log_num_of_rows + count)].value = getuser                                               #日志记载登录用户名
            sheet_log['B' + str(log_num_of_rows + count)].value = today                                                 #日志记载登录日期
            sheet_log['C' + str(log_num_of_rows + count)].value = now                                                   #日志记载登录的精确时间
            sheet_log['D' + str(log_num_of_rows + count)].value = "OK"                                                 #日志记载登录的状态为成功
            workbook.save(filename)                                                                                      # 保存，黑名单和log
            print("welcome")                                                                                           #提示用户成功登录
            stop=False                                                                                                  #修改循环控制变量
            break                                                                                                       #跳出循环
        else:                                                                                                           #密码不正确，提示错误
            print("wrong username or password")                                                                     #提示用户账号或密码错误
            sheet_log['A' + str(log_num_of_rows + count)].value = getuser                                               #日志记载登录用户名
            sheet_log['B' + str(log_num_of_rows + count)].value = today                                                 #日志记载登录日期
            sheet_log['C' + str(log_num_of_rows + count)].value = now                                                   #日志记载登录的精确时间
            sheet_log['D' + str(log_num_of_rows + count)].value = "fail"                                               #日志记载登录的状态为失败
            sheet_log['E' + str(log_num_of_rows + count)].value = "密码错误"                                           #日志记载登录的备注为密码错误
            workbook.save(filename)                                                                                      # 保存，黑名单和log
            loglist.append(getuser)                                                                                      #日志数组增加用户登录名
            if loglist.count(getuser) == 3:                                                                             #日志数组统计，有效的用户3次密码错误
                stop = False                                                                                            #修改循环标志为false，不再循环
                print("You have tried 3 times,goodluck next time")                                                #提示用户，该用户已经尝试3次登录，不能再次尝试
                workbook["blacklist"].append([getuser])                                                                #用户加入黑名单
            else:                                                                                                       #用户不满三次，提示用户是否再次尝试登录
                countine_confirm = input("agein?Y|N:")                                                                 #提示用户选择
                if countine_confirm == "n":                                                                             #用户选择“n”，则退出循环
                    stop = False
    else:                                                                                                               #用户名不在用户列表中，提示错误
        print("wrong username or password")                                                                         #提示用户名或密码错误
        sheet_log['A' + str(log_num_of_rows + count)].value = getuser                                                   #日志记载登录用户名
        sheet_log['B' + str(log_num_of_rows + count)].value = today                                                     #日志记载登录日期
        sheet_log['C' + str(log_num_of_rows + count)].value = now                                                       #日志记载登录的精确时间
        sheet_log['D' + str(log_num_of_rows + count)].value = "fail"                                                   #日志记载登录的状态为失败
        sheet_log['E' + str(log_num_of_rows + count)].value = "用户名错误"                                             #日志记载登录的备注为用户名错误
        workbook.save(filename)                                                                                          # 保存，黑名单和log
        countine_confirm = input("agein?Y|N:")                                                                         #提示用户选择
        if countine_confirm == "n":                                                                                     #用户选择“n”，则退出循环
            stop = False

